import base64
import os
import shutil
import time
from datetime import datetime, timedelta

import openpyxl
from PyQt5.QtGui import QPixmap
from PyQt5.QtWidgets import QDialog

from my_icon import icon_bytes
from utils.my_logger import my_logger as logger


def check_file_exists(filename):
    if os.path.exists(filename):
        print(f"文件已存在: {filename}")
        return True
    else:
        print(f"文件不存在: {filename}")
        return False


def create_folder_if_not_exists(path):
    if not os.path.exists(path):
        print(f"文件路径不存在,创建路径 {path}")
        os.makedirs(path)
    else:
        print(f"文件路径已存在 {path}")


def create_file_if_not_exists(filename):
    if not os.path.exists(filename):
        print(f"文件不存在 {filename}")
        try:
            file = open(filename, 'w')
            file.close()
            print(f"文件创建成功 {filename}")
        except Exception as e:
            print(f"文件创建失败：{filename}" + str(e))
    else:
        print(f"文件已存在: {filename}")


def open_file(path_file):
    try:
        os.startfile(path_file)
        logger.info(f"正在打开文件： {path_file}")
        time.sleep(5)
    except Exception as e:
        logger.error(f"文件打开失败,{str(e)}")


def get_time_now():
    return time.strftime("%Y-%m-%d %H-%M-%S", time.localtime())


def get_current_date():
    now = datetime.now()
    current_date = now.date()
    return current_date


def get_format_date(date):
    return f"{date.year}.{date.month if date.month > 9 else str(date.month)[-1]}.{date.day if date.day > 9 else str(date.day)[-1]}"


def get_format_date_str(date):
    # 直接检查并移除月份和日期前的零
    month_part = date[5:7].lstrip('0') or '0'  # 使用 lstrip 移除前导零，如果结果为空字符串则使用 '0'
    day_part = date[8:].lstrip('0') or '0'  # 同上
    formatted_date_no_leading_zero = f"{date[:4]}.{month_part}.{day_part}"
    return formatted_date_no_leading_zero


def get_yesterday_date():
    today = get_current_date()
    yesterday = today - timedelta(days=1)
    return yesterday


def get_date_days_ago(today, days_ago):
    date = today - timedelta(days=days_ago)
    return date


def get_date_not_year(date):
    format_date = date.strftime('%m%d')
    return format_date


def get_month():
    today = get_current_date()
    month = today.strftime('%Y-%m')
    return month


def get_format_month(month):
    format_month = month.strftime('%Y-%m')
    return format_month


def excel_add_title(file_path, title, merge_cells):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    new_row_data = [title]
    # 获取工作表的最大行数
    max_row = sheet.max_row
    # 遍历每一行，并将其向下移动一行
    for row in range(max_row, 0, -1):
        for col in range(1, sheet.max_column + 1):
            sheet.cell(row=row + 1, column=col, value=sheet.cell(row=row, column=col).value)
        # 在第一行写入新数据
    for col_num, value in enumerate(new_row_data, start=1):
        sheet.cell(row=1, column=col_num, value=value)
    # 合并单元格
    sheet.merge_cells(merge_cells)
    # 保存工作簿
    workbook.save(file_path)


def data_frame_add_title(sheet, title, merge_cells):
    new_row_data = [title]
    # 获取工作表的最大行数
    max_row = sheet.max_row
    # 遍历每一行，并将其向下移动一行
    for row in range(max_row, 0, -1):
        for col in range(1, sheet.max_column + 1):
            sheet.cell(row=row + 1, column=col, value=sheet.cell(row=row, column=col).value)
        # 在第一行写入新数据
    for col_num, value in enumerate(new_row_data, start=1):
        sheet.cell(row=1, column=col_num, value=value)
    # 合并单元格
    sheet.merge_cells(merge_cells)


def copy_and_rename_file(source_path, destination_dir, new_file_name):
    """
    复制文件并重命名。

    参数:
    source_path (str): 源文件的完整路径。
    destination_dir (str): 目标目录的路径。
    new_file_name (str): 新文件的名称（不包含路径）。

    返回:
    str: 目标文件的完整路径（如果复制成功）。
    None: 如果复制失败。
    """
    # 构建目标文件的完整路径
    destination_path = os.path.join(destination_dir, new_file_name)

    # 确保目标目录存在
    os.makedirs(destination_dir, exist_ok=True)

    try:
        # 复制文件并重命名
        shutil.copy2(source_path, destination_path)
        print(f"文件已成功复制到 {destination_path}")
        return destination_path
    except FileNotFoundError:
        print(f"源文件 {source_path} 未找到")
        return None
    except PermissionError:
        print(f"没有权限复制文件到 {destination_dir}")
        return None
    except Exception as e:
        print(f"复制文件时发生错误: {e}")
        return None


# 图标bytes转成pixmap格式
def get_icon():
    icon_img = base64.b64decode(icon_bytes)  # 解码
    icon_pixmap = QPixmap()  # 新建QPixmap对象
    icon_pixmap.loadFromData(icon_img)  # 往QPixmap中写入数据
    return icon_pixmap


if __name__ == '__main__':
    print(get_time_now())