import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill

# 创建简单的数据框
data = {
    '姓名': ['张三', '李四', '王五'],
    '年龄': [23, 34, 45],
    '成绩': [88.5, 92.3, 79.0]
}

df = pd.DataFrame(data)

# 创建一个 ExcelWriter 对象，指定文件名和使用 openpyxl 引擎
with pd.ExcelWriter('styled_table.xlsx', engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name='Sheet1')

    # 加载生成的工作簿
    workbook = writer.book
    worksheet = workbook['Sheet1']

    # 设置样式
    alignment = Alignment(horizontal='center', vertical='center')  # 文本居中
    fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")  # 设置浅黄色背景

    for row in worksheet.iter_rows(min_row=2, max_row=len(df)+1, min_col=1, max_col=len(df.columns)):
        for cell in row:
            cell.alignment = alignment  # 应用居中样式
            cell.fill = fill  # 应用背景色

# 保存文件
print("Excel 文件生成成功!")
